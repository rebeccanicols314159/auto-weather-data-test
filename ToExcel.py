import pandas as pd
import os
import openpyxl as op
from operator import itemgetter

os.chdir(f'{os.getcwd()}/data_files')
rtdir = os.getcwd()
folders = os.listdir()

for i in folders.copy():
    if not "weather_data" in i:
        folders.remove(i)

def getinfo(fders):
    flist = []
    for folder in fders:
        os.chdir(f'{rtdir}/{folder}')
        for file in os.listdir():
            with open(file) as f:
                if 'weather-data' in file:
                    sdate = f.read().replace('\n', ',').split(',')
                    sdate = ['date',folder[13:]] + sdate
                    flist.append(sdate)
    flist = sorted(flist, key=itemgetter(7,0)(itemgetter(3)(flist)))
    return flist

def makedict(info,dates):
    columns = ['observation_time','temp_C', 'windspeedKmph', 'winddirDegree', 'precipMM', 'pressureMB', 'visibilityKm', 'cloudcover']
    infodicts = []
    for o in dates:
        infodict = dict()
        for i in columns:
            entries = []
            for j in info:
                if j[1] == o[13:]: entries.append(j[j.index(i)+1])
            infodict.update({i:tuple((entries))})
        df = pd.DataFrame(infodict)
        infodicts.append(df)
    return infodicts

def makefile(datadicts,sheetnames):
    os.chdir(rtdir)
    wb = op.Workbook()
    ws = wb.active
    ws.title = 'Summary'
    with pd.ExcelWriter('All_data_detailed.xlsx') as writer:
        for i in sheetnames:
            datadicts[0].to_excel(writer, sheet_name=i[13:], index=False)
            datadicts.pop(0)

#Example request
a = getinfo(folders)
b = makedict(a,folders)
makefile(b,folders)